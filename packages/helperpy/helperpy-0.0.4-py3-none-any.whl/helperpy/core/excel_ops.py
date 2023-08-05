from typing import (
    Any,
    ByteString,
    Dict,
    List,
    Optional,
    Union,
)
import io

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import (
    ColumnDimension,
    DimensionHolder,
)
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import win32com.client

from helperpy.core.type_annotations import (
    Number,
    NumberOrString,
)


def style_dataframe_conditionally(
        df_style_obj: Any,
        props: List[Dict[str, Union[pd.Series, List[NumberOrString], Dict[str, str], None]]],
    ) -> Any:
    """
    Takes in Pandas Styler object (of class `pandas.io.formats.style.Styler`), and a list of `props`.
    Returns Pandas Styler object after applying the styles based on the given conditions (via `props`).
    Reference: https://pandas.pydata.org/pandas-docs/stable/user_guide/style.html
    
    In the `props`, each prop is a dictionary having 3 keys: ['row_conditions', 'column_subset', 'styling_kwargs'].
    The keys 'row_conditions' and 'column_subset' can have a value of None.
    The key 'styling_kwargs' can have a value of {}.
    The ordering of the `props` matter i.e; if 5 props are given, the 4th prop can over-write the 3rd prop.

    >>> num_rows = 50
    >>> df = pd.DataFrame(data={
        'column1': [np.random.randint(-25, 25) for _ in range(num_rows)],
        'column2': [np.random.randint(0, 100) for _ in range(num_rows)],
        'column3': [np.random.randint(-40, 15) for _ in range(num_rows)],
        'column4': [np.random.randint(-5, 5) for _ in range(num_rows)],
    })
    >>> props = [
        {
            'row_conditions': (df['column1'] >= 10) | (df['column1'] % 2 == 0),
            'column_subset': ['column1', 'column2'],
            'styling_kwargs': {
                'background-color': 'green',
                'color': 'purple',
            },
        },
        {
            'row_conditions': (df['column3'] < 0),
            'column_subset': None,
            'styling_kwargs': {
                'background-color': 'blue',
                'color': 'orange',
            },
        },
        {
            'row_conditions': None,
            'column_subset': ['column4'],
            'styling_kwargs': {
                'background-color': 'grey',
                'color': 'red',
            },
        },
    ]
    >>> style_dataframe_conditionally(df_style_obj=df.style, props=props)
    """
    df_styled = df_style_obj.__copy__()
    idx = pd.IndexSlice
    for prop in props:
        row_conditions = prop['row_conditions']
        column_subset = prop['column_subset']
        styling_kwargs = prop['styling_kwargs']

        subset = idx[idx[row_conditions], column_subset]
        if row_conditions is None and column_subset is None:
            subset = None
        elif row_conditions is None and column_subset is not None:
            subset = idx[idx[:], column_subset]
        elif row_conditions is not None and column_subset is None:
            subset = idx[idx[row_conditions], :]
        df_styled = df_styled.set_properties(
            subset=subset,
            axis=1,
            **styling_kwargs,
        )
    return df_styled


def style_dataframe_with_background_gradient(
        df_style_obj: Any,
        blue_white_red: Optional[List[NumberOrString]] = None,
        blues: Optional[List[NumberOrString]] = None,
        greens: Optional[List[NumberOrString]] = None,
        red_yellow_green: Optional[List[NumberOrString]] = None,
        summer: Optional[List[NumberOrString]] = None,
    ) -> Any:
    """
    Takes in Pandas Styler object (of class `pandas.io.formats.style.Styler`), and returns the same
    after applying background gradient styles based on the given column subsets.
    """
    df_styled = df_style_obj.__copy__()
    if blue_white_red:
        df_styled = df_styled.background_gradient(subset=blue_white_red, cmap='bwr')
    if blues:
        df_styled = df_styled.background_gradient(subset=blues, cmap='Blues')
    if greens:
        df_styled = df_styled.background_gradient(subset=greens, cmap='Greens')
    if red_yellow_green:
        df_styled = df_styled.background_gradient(subset=red_yellow_green, cmap='RdYlGn')
    if summer:
        df_styled = df_styled.background_gradient(subset=summer, cmap='summer')
    return df_styled


def save_styled_dataframes(
        destination_filepath: str,
        sheet_name_to_df_styled: Dict[str, Any],
    ) -> None:
    """
    Saves Pandas Styler object/s to an Excel file, which could have multiple sheets (one for each Styler object).
    >>> save_styled_dataframes(
        destination_filepath="resulting_file.xlsx",
        sheet_name_to_df_styled={
            "sheet1": df_style_obj_1,
            "sheet2": df_style_obj_2,
        },
    )
    """
    with pd.ExcelWriter(destination_filepath) as excel_writer:
        for sheet_name, df_styled in sheet_name_to_df_styled.items():
            df_styled.to_excel(
                excel_writer=excel_writer,
                sheet_name=sheet_name,
                index=False,
                engine='xlsxwriter',
            )
    return None


def get_all_excel_sheet_names(filepath: str) -> List[str]:
    """Returns list of all sheet names in an Excel file"""
    excel_file = pd.ExcelFile(filepath, engine='openpyxl')
    return excel_file.sheet_names


def get_all_excel_sheets_as_dataframes(filepath: str) -> Dict[str, pd.DataFrame]:
    """
    Takes `filepath` to an Excel file. Returns dictionary having keys = sheet names, and
    values = DataFrame corresponding to respective sheet name.
    """
    sheet_names = get_all_excel_sheet_names(filepath=filepath)
    dict_obj = {}
    for sheet_name in sheet_names:
        df_by_sheet = pd.read_excel(filepath, engine='openpyxl', sheet_name=sheet_name)
        dict_obj[sheet_name] = df_by_sheet
    return dict_obj


def excel_file_to_bytes(filepath: str) -> ByteString:
    """Takes `filepath` to an Excel file, and returns ByteString of the same"""
    bio = io.BytesIO()
    writer = pd.ExcelWriter(bio, engine='openpyxl')
    dict_all_excel_sheets = get_all_excel_sheets_as_dataframes(filepath=filepath)
    for sheet_name, df_obj in dict_all_excel_sheets.items():
        df_obj.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()
    bytes_obj = bio.getvalue()
    return bytes_obj


def __get_column_names_from_worksheet(worksheet: Worksheet) -> List[NumberOrString]:
    """Returns list of column names (headers) from worksheet object (i.e; the values from row #1)"""
    column_names = [
        column.value for column in next(worksheet.iter_rows(min_row=1, max_row=1))
    ]
    return column_names


def adjust_column_widths_of_excel_file(
        src_filepath: str,
        destination_filepath: str,
        how: Optional[str] = 'column_header_and_values',
        fixed_width: Optional[Number] = None,
        sheets_subset: Optional[List[str]] = None,
    ) -> None:
    """
    Adjusts the widths of the columns in the Excel file given.
    Works only for non-hierarchical columns (flat columns).
    Note: Only works for Excel files with the 'xlsx' extension.

    Parameters:
        - src_filepath (str): Path to the source Excel file (along with the extension).
        - destination_filepath (str): Path to the destination Excel file (along with the extension).
        - how (str): How to assign the column widths. Options: ['column_header', 'column_header_and_values', 'fixed']. Default: 'column_header_and_values'.
        If `how` is set to 'column_header', the column widths will be the size of the column header.
        If `how` is set to 'column_header_and_values', the column widths will be the size of the longest of the column's header and values.
        If `how` is set to 'fixed', please set the `fixed_width` parameter.
        - fixed_width (int | float): Fixed width to use for the columns. To be set only if `how` is set to 'fixed'. Default: None.
        - sheets_subset (list): Subset of sheets for which the column widths need to be altered. Default: None.
    
    >>> adjust_column_widths_of_excel_file(
            src_filepath="data.xlsx",
            destination_filepath="data_with_adjusted_columns.xlsx",
            how='fixed',
            fixed_width=15,
            sheets_subset=['Sheet1', 'Sheet2', 'Sheet5'],
        )
    """
    how_options = ['column_header', 'column_header_and_values', 'fixed']
    if how not in how_options:
        raise ValueError(f"Expected `how` to be in {how_options}, but got '{how}'")
    if how == 'fixed' and fixed_width is None:
        raise ValueError("Since you have set how='fixed', please set the `fixed_width` parameter as well")
    
    workbook = openpyxl.load_workbook(filename=src_filepath)
    sheet_names = workbook.sheetnames if sheets_subset is None else sheets_subset
    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        column_names = __get_column_names_from_worksheet(worksheet=worksheet)
        dimension_holder = DimensionHolder(worksheet=worksheet)
        for column_idx, column_name in enumerate(column_names):
            column_idx += 1 # Since 1-based indexing is used
            column_letter = get_column_letter(idx=column_idx)
            if how == 'column_header':
                width = len(column_name) * 1.2
            elif how == 'column_header_and_values':
                cell_objs_by_column = worksheet[column_letter]
                width = max(map(lambda cell_obj: len(str(cell_obj.value)), cell_objs_by_column)) * 1.2
            elif how == 'fixed':
                width = fixed_width
            dimension_holder[column_letter] = ColumnDimension(
                worksheet=worksheet,
                min=column_idx,
                max=column_idx,
                width=width,
            )
        worksheet.column_dimensions = dimension_holder
    workbook.save(filename=destination_filepath)
    return None


def excel_to_pdf(
        src_filepath: str,
        destination_pdf_filepath: str,
        sheets_subset: Optional[List[str]] = None,
    ) -> None:
    r"""
    Takes `src_filepath` to an Excel file, and converts it to PDF file and saves it to `destination_pdf_filepath`.
    Note: Requires absolute filepaths. Does not work with relative filepaths.
    Reference: https://stackoverflow.com/questions/66421969/how-to-convert-excel-to-pdf-using-python

    >>> excel_to_pdf(
            src_filepath=r"C:\files\excel_files\some_data.xlsx",
            destination_pdf_filepath=r"C:\files\pdf_files\some_data.pdf",
            sheets_subset=['Sheet1', 'Sheet2', 'Sheet5'],
        )
    """
    sheet_names = get_all_excel_sheet_names(filepath=src_filepath)
    sheet_indices = [idx+1 for idx, _ in enumerate(sheet_names)] # Set 1-based indices
    if sheets_subset is not None:
        sheets_that_do_not_exist = list(
            set(sheets_subset).difference(set(sheet_names))
        )
        if sheets_that_do_not_exist:
            raise ValueError(f"The following sheet names do not exist: {sheets_that_do_not_exist}. Please correct the `sheets_subset` parameter")
        for sheet_idx, sheet_name in zip(sheet_indices, sheet_names):
            if sheet_name not in sheets_subset:
                sheet_indices.remove(sheet_idx)
                sheet_names.remove(sheet_name)
    dispatch_obj = win32com.client.Dispatch(dispatch="Excel.Application")
    dispatch_obj.Visible = False
    wb_obj = dispatch_obj.Workbooks.Open(src_filepath)
    wb_obj.WorkSheets(sheet_indices).Select()
    wb_obj.ActiveSheet.ExportAsFixedFormat(0, destination_pdf_filepath)
    wb_obj.Close()
    dispatch_obj.Quit()
    return None