from importlib.util import find_spec
from functools import wraps
from pathlib import Path
from typing import Collection, Any, Type, List, TypeVar, Iterator

# Conditional imports
if find_spec('openpyxl'):
    import openpyxl
    from openpyxl.cell import WriteOnlyCell
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.cell.cell import Cell, TYPE_ERROR
else:
    openpyxl = None
    WriteOnlyCell = None
    Cell = None
    TYPE_ERROR = 'e'
    Workbook = TypeVar('Workbook')
    Worksheet = TypeVar('Worksheet')

from ..normalization import normalize_row

from ..decorators import requires
from ..spreadsheets import SpreadsheetSerializer
from ..typing import ParsedDataRow


__ALL__ = ['XLSXSerializer', 'dump', 'write_table', 'load', 'read_table', 'read_headers', 'read_schema', 'iter_table',
           'iter_structure']


class XLSXSerializer(SpreadsheetSerializer):
    MAX_LINES = 1048576
    MAX_COLS = 16384

    # OpenPyXL specific implementation
    @staticmethod
    @requires('openpyxl')
    def _add_sheet(workbook: Workbook, sheet_name: str) -> Worksheet:
        return workbook.create_sheet(title=sheet_name)

    @staticmethod
    @requires('openpyxl')
    def _write_headers(sheet: Worksheet, headers: Collection[str]):
        sheet.append((WriteOnlyCell(sheet, h) for h in headers))

    @staticmethod
    @requires('openpyxl')
    def _write_row(sheet: Worksheet, headers: Collection[str], row_num: int, row_data: ParsedDataRow,
                   default: Any = SpreadsheetSerializer.DEFAULT_EMPTY_ON_WRITE):
        sheet.append(list(normalize_row(row_data, headers, default=default).values()))

    @staticmethod
    @requires('openpyxl')
    def _freeze_headers(sheet: Worksheet):
        # TODO: This probably requires not using a write-only sheet
        # sheet.freeze_panes = sheet.cell(2, 1)
        pass

    @staticmethod
    @requires('openpyxl')
    def _get_new_workbook(**new_kwargs) -> Workbook:
        new_kwargs.setdefault('write_only', True)
        return openpyxl.Workbook(**new_kwargs)

    @staticmethod
    @requires('openpyxl')
    def _save_workbook(workbook: Workbook, fp: Path | str):
        workbook.save(fp)

    @requires('openpyxl')
    def _iter_through_sheet(self,
                            sheet: Worksheet,
                            workbook: Workbook,
                            constructor: Type = None,
                            default: Any = SpreadsheetSerializer.DEFAULT_EMPTY_ON_READ,
                            ) -> Iterator[ParsedDataRow]:
        headers = None

        for row in sheet.rows:

            # Extract headers from first row
            if headers is None:
                headers = [c.value for c in row]
                continue

            row = dict(zip(headers, (self._parse_cell_value(c, default) for c in row)))

            yield constructor(**row) if constructor is not None else row

    @requires('openpyxl')
    def _parse_cell_value(self, cell: Cell, default: Any = SpreadsheetSerializer.DEFAULT_EMPTY_ON_READ) -> Any:
        value = cell.value
        if cell.data_type == TYPE_ERROR:
            return self.CellError.match_behaviour(self.cell_error, value, default)

        return value

    @staticmethod
    @requires('openpyxl')
    def _get_sheet_name(sheet: Worksheet) -> str:
        return sheet.title

    @staticmethod
    @requires('openpyxl')
    def _get_all_sheets(book: Workbook) -> List[Any]:
        return book.worksheets

    @staticmethod
    @requires('openpyxl')
    def _open_workbook(fp: Path | str, for_append=False, **kwargs) -> Workbook:
        return openpyxl.load_workbook(fp, read_only=not for_append, data_only=not for_append, **kwargs)


# Default implementation
__xlsx_serializer = XLSXSerializer()

write_table = wraps(__xlsx_serializer.write_table)
dump = wraps(__xlsx_serializer.dump)
load = wraps(__xlsx_serializer.load)
read_table = wraps(__xlsx_serializer.read_table)
read_headers = wraps(__xlsx_serializer.read_headers)
read_schema = wraps(__xlsx_serializer.read_schema)
iter_table = wraps(__xlsx_serializer.iter_table)
iter_structure = wraps(__xlsx_serializer.iter_structure)
