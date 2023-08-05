# cython: language_level=3
from typing import Callable as _Callable, Tuple as _Tuple, Union as _Union, Iterable as _Iterable, Optional as _Optional
from threading import Thread as _Thread, Lock as _Lock, Event as _Event
from datetime import datetime as _datetime
from time import time as _now
import sys as _sys
import traceback as _traceback
from openpyxl import load_workbook as _load_xlsx
from typing import List as _List, Dict as _Dict
from watchdog.observers import Observer as _FSObserver
from watchdog.events import FileSystemEventHandler as _FileSystemEventHandler


class FileSystemObserver(_FileSystemEventHandler):
	def __init__(self, path: str, recursive: bool = True):
		self.__observer = _FSObserver()
		self.__observer.schedule(self, path, recursive)
		self.__observer.start()
		self.__changed = False

	@property
	def changed(self) -> bool:
		return self.__changed

	def reset(self):
		self.__changed = False

	def on_any_event(self, event):
		self.__changed = True

	def close(self):
		if self.__observer is not None:
			self.__observer.stop()
			self.__observer.join()
			self.__observer = None

	def __del__(self):
		self.close()


def int_2_bool_list(integer, bit_count) -> _List[bool]:
	result = [False] * bit_count
	for i in range(bit_count):
		if (integer & (1 << i)) != 0:
			result[i] = True
	return result


def bool_list_2_int(bools: _List[bool]) -> int:
	result = 0
	for idx, b in enumerate(bools):
		if b:
			result |= (1 << idx)

	return result


class Callback:
	def __init__(self):
		self._callback_list = list()
		self._callback_list_guard = _Lock()

	def subscribe(self, callback: _Callable):
		with self._callback_list_guard:
			if callback not in self._callback_list:
				self._callback_list.append(callback)

	def unsubscribe(self, callback: _Callable):
		with self._callback_list_guard:
			if callback in self._callback_list:
				self._callback_list.remove(callback)

	def unsubscribe_all(self):
		with self._callback_list_guard:
			self._callback_list.clear()

	def __call__(self, *args, **kwargs):
		results = dict()
		with self._callback_list_guard:
			for callback in self._callback_list:
				results[callback] = callback(*args, **kwargs)
		return results


def timestamp(stamp: float = None) -> str:
	if stamp is None:
		stamp = _now()
	dt = _datetime.fromtimestamp(stamp)
	return dt.isoformat(' ')[:19]


def timestamp_as_filename(stamp: float = None, sep: str = '-') -> str:
	return timestamp(stamp).replace(':', sep)


def format_exception(skip_level: int = 0):
	exc_type, exc_value, exc_tb = _sys.exc_info()
	for i in range(skip_level):
		if exc_tb.tb_next is not None:
			exc_tb = exc_tb.tb_next

	return _traceback.format_exception(exc_type, exc_value, exc_tb)


def print_exception(skip_level: int = 0):
	lines = format_exception(skip_level)
	print(''.join(lines), end='')


class StrState:
	__states__: _List[str] = list()
	__strict__ = True

	@classmethod
	def all(cls) -> _Tuple[str]:
		return tuple(cls.__states__)

	@classmethod
	def add_state(cls, value: str):
		assert isinstance(value, str)
		assert value not in cls.__states__
		cls.__states__.append(value)

	def __init__(self, value: _Union[str]):
		if isinstance(value, str):
			assert value in self.__states__
			self.__value = value
		else:
			assert isinstance(value, StrState)
			self.__value = value.__value

	def set(self, value: _Union[str]):
		if isinstance(value, str):
			assert value in self.__states__
			self.__value = value
		else:
			assert isinstance(value, StrState)
			self.__value = value.__value

	def __str__(self):
		return self.__value

	def __eq__(self, other):
		if isinstance(other, str):
			if self.__strict__:
				assert other in self.__states__
			return self.__value == other
		assert isinstance(other, StrState)
		return self.__value == other.__value

	def __int__(self):
		return self.__states__.index(self.__value)

	def __lt__(self, other):
		if isinstance(other, StrState):
			return int(self) < int(other)
		else:
			assert isinstance(other, str)
			return int(self) < self.__states__.index(other)


def load_a_list_of_dict_from_xlsx(
		xlsx_name: str,
		sheet: _Union[int, str] = 0,
		column_count: int = 0,
		row_count: int = 0
) -> _List[dict]:
	book = _load_xlsx(xlsx_name, read_only=True, data_only=True)
	if isinstance(sheet, int):
		page = book.worksheets[sheet]
	else:
		page = book.get_sheet_by_name(sheet)
	headers = []
	column_index = 1
	while column_index <= column_count or column_count <= 0:
		header = page.cell(1, column_index).value
		if header is None or header == '':
			column_count = column_index - 1
			break
		else:
			headers.append(header)
			column_index += 1
	data_rows = []
	row_index = 2
	while row_count == 0 or (row_index - 1) <= row_count:
		data_row = []
		if page.cell(row_index, 1).value in [None, '']:
			break
		for _column_idx in range(column_count):
			column_index = _column_idx + 1
			value = page.cell(row_index, column_index).value
			data_row.append(value)
		data_rows.append(data_row)
		row_index += 1

	result = []
	for data_row in data_rows:
		item = dict()
		for index in range(column_count):
			item[headers[index]] = data_row[index]
		result.append(item)

	return result


def load_a_dict_of_dict_from_xlsx(
		xlsx_name: str,
		sheet: _Union[int, str] = 0,
		column_count: int = 0,
		row_count: int = 0,
		key: _Union[int, str] = 0
) -> _Dict[str, dict]:
	book = _load_xlsx(xlsx_name, read_only=True, data_only=True)
	if isinstance(sheet, int):
		page = book.worksheets[sheet]
	else:
		page = book.get_sheet_by_name(sheet)
	headers = []
	column_index = 1
	while column_index <= column_count or column_count <= 0:
		header = page.cell(1, column_index).value
		if header is None or header == '':
			column_count = column_index - 1
			break
		else:
			headers.append(header)
			column_index += 1

	if isinstance(key, int):
		key_index = key
	elif isinstance(key, str):
		key_index = headers.index(key)
	else:
		raise TypeError('Argument "key" must be an object of int or str.')

	data_rows = []
	row_index = 2
	while row_count == 0 or (row_index - 1) <= row_count:
		data_row = []
		if page.cell(row_index, 1).value in [None, '']:
			break
		for _column_idx in range(column_count):
			column_index = _column_idx + 1
			value = page.cell(row_index, column_index).value
			data_row.append(value)
		data_rows.append(data_row)
		row_index += 1

	result = dict()
	for data_row in data_rows:
		item = dict()
		key_value = data_row[key_index]
		if key_value is None:
			raise ValueError('None key found.')
		for index in range(column_count):
			item[headers[index]] = data_row[index]
		result[key_value] = item

	return result


def get_the_best_value(lower: float, upper: float, values: _Iterable) -> _Optional[float]:
	"""
	在一系列数据（values）中，找到离给定的两个数（lower and upper）的距离之和最小的那个数。
	这个函数已经考虑了“inf”和“-inf”的情况，希望能满足你的需求。不过，如果提供的数据中确实需要考虑正负无穷的情况，建议你还是先做一下测试，
	看是否跟你期望的行为一致。
	数据中包含的“NAN”值将会被忽略。
	"""
	shift = float('inf')
	best = None
	for value in values:
		distance = abs((upper - value) - (value - lower))
		if distance < shift:
			shift = distance
			best = value
		elif best is None and distance == shift:
			best = value
	return best


def get_middle_value(values: _Iterable, in_place: bool = False) -> _Optional[float]:
	"""
	返回一系列数值中的中位数。
	如果这系列数值中没有值，返回 None；
	如果具有偶数个数值，返回中间两个值的平均值；
	如果具有奇数个值，返回中间那个值
	"""
	if in_place and hasattr(values, 'sort'):
		sorted_values = values
	else:
		sorted_values = list(values)
	sorted_values.sort()
	count = len(sorted_values)
	if count == 0:
		return None
	index = count // 2
	if count % 2 == 0:
		return (sorted_values[index - 1] + sorted_values[index]) / 2
	return sorted_values[index]


def make_filename_valid(filename: str, char_map: _Optional[_Dict[str, str]] = None, default: str = '') -> str:
	"""
	文件系统的文件命名规则里，规定有一些字符不能用于文件名。而在自动化生成文件名时，文件名来源可能会包含有这些字符。
	为了顺利生成文件，我们需要把这些非法字符去掉或替换掉。这就是这个函数的功能。
	"""
	if char_map is None:
		char_map = dict()

	for char in '\\/:*?"<>|':
		replacement = char_map.get(char, default)
		filename = filename.replace(char, replacement)
	return filename


class ThreadProxy:
	"""
	Usage example:
	def a_long_time_function(arg):
		# do something
		return 0
	proxy = ThreadProxy(True, a_long_time_function, 'This is the arg')
	# do something while function a_long_time_function is running.
	result = proxy.result  # This will wait until a_long_time_function finish
	# do something with result
	"""
	def __init__(self, startup_at_once: bool, func: _Callable, *args, **kwargs):
		self.__func = func
		self.__args = args
		self.__kwargs = kwargs
		self.__ready = False
		self.__event = _Event()
		self.__result = None
		self.__thread: _Optional[_Thread] = None
		self.__exception: _Optional[BaseException] = None

		if startup_at_once:
			self.start()

	def start(self):
		if self.__thread is not None:
			raise ValueError('Thread has been started. No starting again.')
		self.__thread = _Thread(target=self.__call_at_backstage)
		self.__thread.start()

	def __call_at_backstage(self):
		try:
			result = self.__func(*self.__args, **self.__kwargs)
		except BaseException as e:
			self.__exception = e
		else:
			self.__result = result

		self.__ready = True
		self.__event.set()

	@property
	def ready(self) -> bool:
		return self.__ready

	def __check_and_return(self):
		if self.__exception is not None:
			raise self.__exception
		if self.__thread is not None:
			self.__thread.join()
			self.__thread = None
		return self.__result

	@property
	def result(self):
		if self.__thread is None and not self.ready:
			raise ValueError('Thread has not been started. You must call "start" first.')
		self.__event.wait()
		return self.__check_and_return()

	def get(self, timeout: float):
		if self.__thread is None and not self.ready:
			raise ValueError('Thread has not been started. You must call "start" first.')
		if not self.__event.wait(timeout):
			raise TimeoutError('Wait for thread proxy timeout!')
		return self.__check_and_return()
